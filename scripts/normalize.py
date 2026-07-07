#!/usr/bin/env python
"""
normalize.py — Этап 2: нормализация xlsx-отчётов в длинную таблицу.

Логика колонок унаследована от process_admissions_25.py (сопоставление по
регэксп-сигнатурам, а не по позиции), адаптирована под шапку этого года:

- Источник — xlsx (не CSV), читаем через openpyxl.
- Заголовок таблицы — не на фиксированной строке: ищем строку, где первая
  ячейка равна "№ п/п".
- Сразу после заголовка идёт служебная строка-водяной знак вида
  "~36913483045~" — отсекаем её (и все прочие "хвостовые" пустые строки)
  фильтром по тому, что "№ п/п" должно быть числом.
- Между колонкой приоритета и "Суммой конкурсных баллов" стоит ПЕРЕМЕННОЕ
  число (1 или 2) колонок вступительных испытаний с произвольным названием
  программы (не только "Конкурс портфолио" — бывает "Математика",
  "Собеседование", "Экзамен по ..." и т.д.). Эти колонки не сопоставляются
  по имени, а определяются позиционно и раскладываются в универсальные
  слоты entry_test_1 / entry_test_2.
- Campus/program_name берутся из проверенного на Этапе 0 списка целей
  (data/stage0_targets_*.csv), а не только из ячейки A1 — так результат
  остаётся согласован с тем, что вы уже сверяли глазами.
"""
from __future__ import annotations

import argparse
import csv
import glob
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

RE_FILENAME = re.compile(r"(\d+)_(Budget|Commercial)\.xlsx$")

LONG_TABLE_COLUMNS = [
    "student_id", "reg_number", "program_id", "place_type", "priority",
    "target_quota_priority", "is_target_quota_applicant",
    "entry_test_1_name", "entry_test_1_score",
    "entry_test_2_name", "entry_test_2_score",
    "total_score", "total_score_quota", "all_grades_positive",
    "competition_status", "enrollment_consent", "docs_returned",
    "education_doc_type", "has_contract", "first_payment_made",
    "orig_doc_submitted",
]


def normalize_header(h) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h)).strip().lower()


def load_program_lookup(targets_csv: Path) -> dict[str, dict]:
    lookup: dict[str, dict] = {}
    with targets_csv.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            lookup[row["program_id"]] = {
                "program_name": row["program_name"],
                "campus": row["campus"],
            }
    return lookup


def find_header_row(ws) -> tuple[int, list]:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row and row[0] == "№ п/п":
            header = [h for h in row if h is not None]
            return i, header
    raise ValueError("Не найдена строка заголовка ('№ п/п' в первых 10 строках)")


def locate_columns(header: list) -> dict:
    norm = [normalize_header(h) for h in header]

    def find(pattern: str, exclude: str | None = None) -> int | None:
        for i, h in enumerate(norm):
            if re.search(pattern, h) and (exclude is None or not re.search(exclude, h)):
                return i
        return None

    total_score_idx = find(r"сумма.*конкурсн.*балл", exclude=r"квот")
    if total_score_idx is None:
        raise ValueError("Не найдена колонка 'Сумма конкурсных баллов'")

    priority_idx = None
    for i in range(total_score_idx - 1, -1, -1):
        if "приоритет" in norm[i] and "мест" in norm[i]:
            priority_idx = i
            break
    if priority_idx is None:
        raise ValueError("Не найдена колонка приоритета перед 'Сумма конкурсных баллов'")

    entry_test_idx = list(range(priority_idx + 1, total_score_idx))

    cols = {
        "student_id": find(r"уникальн.*код.*поступ"),
        "reg_number": find(r"рег\.?\s*номер|регистрацион.*номер"),
        "priority_budget": find(r"приоритет.*бюджетн.*мест"),
        "priority_commercial": find(r"приоритет.*коммерч.*мест"),
        "priority_target": find(r"приоритет.*целев.*мест"),
        "is_target_quota": find(r"поступление.*целев.*квот"),
        "total_score": total_score_idx,
        "total_score_quota": find(r"сумма.*конкурсн.*балл.*квот"),
        "all_grades_positive": find(r"все.*оценк.*положительн"),
        "competition_status": find(r"статус.*участ.*конкурс"),
        "enrollment_consent": find(r"согласие.*зачисл"),
        "docs_returned": find(r"возврат.*документ"),
        "education_doc_type": find(r"вид.*документ.*образован"),
        "has_contract": find(r"наличие.*договор.*образован"),
        "first_payment_made": find(r"оплата.*перв.*период.*договор"),
        "orig_doc_submitted": find(r"дата.*предоставлен.*подлинник"),
        "entry_test_idx": entry_test_idx,
    }
    return cols


def load_one_file(path: Path, program_lookup: dict) -> pd.DataFrame:
    m = RE_FILENAME.search(path.name)
    if not m:
        raise ValueError(f"Имя файла не соответствует шаблону {{ID}}_{{Тип}}.xlsx: {path.name}")
    program_id, place_type_raw = m.groups()
    place_type = "budget" if place_type_raw == "Budget" else "commercial"

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row_idx, header = find_header_row(ws)
    cols = locate_columns(header)

    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        row = list(row[: len(header)])
        n_pp = row[0]
        if not isinstance(n_pp, (int, float)):
            continue  # водяной знак или хвостовая пустая строка

        def get(key):
            idx = cols.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        entry_tests = [(header[i], row[i]) for i in cols["entry_test_idx"] if i < len(row)]
        entry_tests += [(None, None)] * (2 - len(entry_tests))  # pad to 2 slots

        priority = get("priority_budget") if place_type == "budget" else get("priority_commercial")

        rec = {
            "student_id": get("student_id"),
            "reg_number": get("reg_number"),
            "program_id": program_id,
            "place_type": place_type,
            "priority": priority,
            "target_quota_priority": get("priority_target"),
            "is_target_quota_applicant": get("is_target_quota"),
            "entry_test_1_name": entry_tests[0][0],
            "entry_test_1_score": entry_tests[0][1],
            "entry_test_2_name": entry_tests[1][0],
            "entry_test_2_score": entry_tests[1][1],
            "total_score": get("total_score"),
            "total_score_quota": get("total_score_quota"),
            "all_grades_positive": get("all_grades_positive"),
            "competition_status": get("competition_status"),
            "enrollment_consent": get("enrollment_consent"),
            "docs_returned": get("docs_returned"),
            "education_doc_type": get("education_doc_type"),
            "has_contract": get("has_contract"),
            "first_payment_made": get("first_payment_made"),
            "orig_doc_submitted": get("orig_doc_submitted"),
        }
        records.append(rec)
    wb.close()

    df = pd.DataFrame.from_records(records, columns=LONG_TABLE_COLUMNS)

    prog_info = program_lookup.get(program_id)
    if prog_info is None:
        print(f"⚠ {path.name}: program_id {program_id} отсутствует в списке Этапа 0", file=sys.stderr)
        df["program_name"] = None
        df["campus"] = None
    else:
        df["program_name"] = prog_info["program_name"]
        df["campus"] = prog_info["campus"]

    df["source_file"] = path.name

    for c in ("priority", "target_quota_priority"):
        df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")
    for c in ("total_score", "total_score_quota", "entry_test_1_score", "entry_test_2_score"):
        df[c] = pd.to_numeric(df[c], errors="coerce")

    if len(df) > 0 and df["student_id"].isna().all():
        raise ValueError(f"{path.name}: не удалось извлечь ни одного 'Уникальный код поступающего'")

    return df


def main():
    p = argparse.ArgumentParser(description="Этап 2: нормализация xlsx в длинную таблицу")
    p.add_argument("raw_dir", type=Path, help="папка с xlsx (например raw/2026-07-03)")
    p.add_argument("targets_csv", type=Path, help="CSV со списком программ из parse_toc.py")
    p.add_argument("-o", "--out", type=Path, required=True, help="куда сохранить итог (.parquet)")
    args = p.parse_args()

    program_lookup = load_program_lookup(args.targets_csv)
    files = sorted(args.raw_dir.glob("*.xlsx"))
    if not files:
        print(f"В {args.raw_dir} не найдено xlsx-файлов", file=sys.stderr)
        sys.exit(1)

    parts = []
    n_ok, n_err = 0, 0
    for fp in files:
        try:
            parts.append(load_one_file(fp, program_lookup))
            n_ok += 1
        except Exception as e:
            print(f"⚠ {fp.name}: {e}", file=sys.stderr)
            n_err += 1

    if not parts:
        print("Не удалось обработать ни одного файла", file=sys.stderr)
        sys.exit(1)

    long_df = pd.concat(parts, ignore_index=True)
    long_df["student_id"] = long_df["student_id"].astype("Int64")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    long_df.to_parquet(args.out, index=False)

    print(f"Обработано файлов: ok={n_ok} err={n_err}")
    print(f"Строк в длинной таблице: {len(long_df)}")
    print(f"Уникальных абитуриентов: {long_df['student_id'].nunique()}")
    print(f"Уникальных программ: {long_df['program_id'].nunique()}")
    print(f"Сохранено: {args.out}")


if __name__ == "__main__":
    main()
