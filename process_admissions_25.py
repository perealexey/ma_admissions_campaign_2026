#!/usr/bin/env python
"""
process_admissions.py — объединяет выгрузки списков абитуриентов (CSV) в сводную таблицу.

**Правка 2025‑08‑07:** добавлена поддержка бюджетных и коммерческих мест.
Скрипт обрабатывает файлы из папок budget и commercial, создавая сводную таблицу
с тремя колонками на программу: приоритет бюджетного места, приоритет коммерческого места, баллы.
"""
from __future__ import annotations
import argparse, csv, re, sys, textwrap, zipfile
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import pandas as pd

RE_PROGRAM = re.compile(r"Образовательная программа\s+\"+(.+?)\"+$")

# Ключевые колонки и сигнатуры для поиска (регистронезависимо, пробелы игнорируются)
COL_SIGNATURES = {
    "Регистрационный номер": [r"регистрацион.*номер"],
    "Приоритет бюджетного места": [r"приоритет.*бюджетн.*мест"],
    "Приоритет коммерческого места": [r"приоритет.*коммерч.*мест"],
    "Сумма конкурсных баллов": [r"сумма.*конкурсн.*балл(?!.*квот)"]
}
NEEDED_COLS = list(COL_SIGNATURES)

# ―――――――――― Вспомогательные ――――――――――

def normalize(col: str) -> str:
    """Сводим название колонки к единому виду: латиница → lower, \s+ → ' '."""
    return re.sub(r"\s+", " ", col).strip().lower()


def detect_encoding(path: Path) -> str:
    try:
        with path.open("r", encoding="utf-8-sig") as _:
            return "utf-8-sig"
    except UnicodeDecodeError:
        return "cp1251"


def read_head(path: Path, n: int = 15) -> List[str]:
    enc = detect_encoding(path)
    with path.open("r", encoding=enc, errors="ignore") as fp:
        return [next(fp, "") for _ in range(n)]


def extract_program(lines: List[str]) -> str:
    for line in lines:
        m = RE_PROGRAM.search(line.strip())
        if m:
            return m.group(1).strip()
    return "(неизвестная программа)"


def find_table_start(lines: List[str]) -> int:
    for idx, line in enumerate(lines):
        # Check for both lines containing numbers and headers
        if "№ п/п" in line and "Регистрационный номер" in line:
            return idx
    return 8  # fallback


def unify_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapping: Dict[str, str] = {}
    for col in df.columns:
        norm = normalize(col)
        for target, patterns in COL_SIGNATURES.items():
            if any(re.search(p, norm) for p in patterns):
                mapping[col] = target
                break
    df = df.rename(columns=mapping)
    for tgt in NEEDED_COLS:
        if tgt not in df.columns:
            df[tgt] = pd.NA
    return df[NEEDED_COLS]


def load_single(path: Path, place_type: str) -> pd.DataFrame:
    """Загружает один CSV файл, определяя тип места (budget/commercial)."""
    head = read_head(path, 20)  # Read more lines to find headers
    programme = extract_program(head)
    skiprows = find_table_start(head)
    enc = detect_encoding(path)

    df = pd.read_csv(
        path,
        sep=";",
        engine="python",
        skiprows=skiprows,
        encoding=enc,
        quoting=csv.QUOTE_MINIMAL,
        dtype=str,
    ).dropna(axis=0, how="all").dropna(axis=1, how="all")

    # If the columns don't look right, try to find the correct header row
    if (len(df.columns) < 5 or 
        not any("регистрацион" in str(col).lower() and "номер" in str(col).lower() for col in df.columns)):
        
        # Try to find the correct header row by scanning for "Регистрационный номер"
        for i in range(15):
            try:
                test_df = pd.read_csv(path, sep=";", engine="python", skiprows=i, encoding=enc, 
                                    quoting=csv.QUOTE_MINIMAL, dtype=str, nrows=1)
                test_df = test_df.dropna(axis=1, how="all")
                
                if any("регистрацион" in str(col).lower() and "номер" in str(col).lower() for col in test_df.columns):
                    df = pd.read_csv(path, sep=";", engine="python", skiprows=i, encoding=enc,
                                   quoting=csv.QUOTE_MINIMAL, dtype=str).dropna(axis=0, how="all").dropna(axis=1, how="all")
                    break
            except:
                continue
        else:
            raise ValueError(f"Не удалось найти правильные заголовки в {path.name}")
    
    # Check if we have the key columns before unification
    has_reg_num = any("регистрацион" in str(col).lower() and "номер" in str(col).lower() for col in df.columns)
    if not has_reg_num:
        raise ValueError(f"Не найдена колонка 'Регистрационный номер'. Доступные колонки: {list(df.columns)[:5]}")

    df = unify_columns(df)
    
    # Определяем какой приоритет есть в файле
    priority_col = None
    if place_type == "budget":
        priority_col = "Приоритет бюджетного места"
    elif place_type == "commercial":
        priority_col = "Приоритет коммерческого места"
    
    # Преобразуем колонки в числовые
    if priority_col and priority_col in df.columns:
        df[priority_col] = pd.to_numeric(df[priority_col], errors="coerce")
    df["Сумма конкурсных баллов"] = pd.to_numeric(df["Сумма конкурсных баллов"], errors="coerce")
    
    df["Программа"] = programme
    df["Тип места"] = place_type

    if df["Регистрационный номер"].notna().sum() == 0:
        raise ValueError("Не удалось извлечь регистрационные номера — проверьте CSV")
    return df


def iter_csv_files(src: Path):
    if src.is_dir():
        yield from sorted(src.glob("*.csv"))
    elif src.suffix.lower() == ".zip":
        with zipfile.ZipFile(src) as zf:
            for zi in zf.infolist():
                # Skip macOS hidden files and directories
                filename = zi.filename
                if (filename.lower().endswith(".csv") and 
                    not filename.startswith("._") and 
                    not filename.startswith("__MACOSX/")):
                    with zf.open(zi) as f:
                        tmp = Path("/tmp") / Path(zi.filename).name
                        tmp.write_bytes(f.read())
                    yield tmp
    else:
        raise ValueError("Ожидается папка или ZIP с CSV")


def build_combined_from_folders(budget_dir: Path, commercial_dir: Path):
    """Обрабатывает файлы из папок budget и commercial, создавая единую сводную таблицу."""
    parts = []
    budget_count = 0
    commercial_count = 0
    
    # Обрабатываем бюджетные файлы
    print("Обработка бюджетных файлов:")
    if budget_dir.exists():
        for csv_p in iter_csv_files(budget_dir):
            try:
                part = load_single(csv_p, "budget")
                parts.append(part)
                budget_count += 1
                print(f"✓ {csv_p.name}: {len(part)} строк (бюджет)")
            except Exception as e:
                print(f"⚠ {csv_p.name}: {e}", file=sys.stderr)
    else:
        print(f"⚠ Папка {budget_dir} не найдена")
    
    # Обрабатываем коммерческие файлы
    print("\nОбработка коммерческих файлов:")
    if commercial_dir.exists():
        for csv_p in iter_csv_files(commercial_dir):
            try:
                part = load_single(csv_p, "commercial")
                parts.append(part)
                commercial_count += 1
                print(f"✓ {csv_p.name}: {len(part)} строк (коммерция)")
            except Exception as e:
                print(f"⚠ {csv_p.name}: {e}", file=sys.stderr)
    else:
        print(f"⚠ Папка {commercial_dir} не найдена")
    
    if not parts:
        raise RuntimeError("Нет корректных CSV для обработки")

    # Объединяем все данные
    long_df = pd.concat(parts, ignore_index=True)
    
    # Создаем расширенный pivot с тремя колонками на программу
    wide_df = create_three_column_pivot(long_df)
    
    print(f"\nОбработано:")
    print(f"•  {budget_count} файлов из папки budget")
    print(f"•  {commercial_count} файлов из папки commercial")
    print(f"•  Всего {budget_count + commercial_count} файлов")
    
    return long_df, wide_df


def create_three_column_pivot(long_df: pd.DataFrame) -> pd.DataFrame:
    """Создает pivot таблицу с тремя колонками на программу."""
    
    # Приводим регистрационные номера к строковому типу
    long_df["Регистрационный номер"] = long_df["Регистрационный номер"].astype(str)
    
    # Получаем все уникальные программы
    programs = sorted(long_df["Программа"].unique())
    reg_numbers = sorted(long_df["Регистрационный номер"].unique())
    
    # Создаем структуру для результата
    result_data = []
    
    for reg_num in reg_numbers:
        row_data = {"Регистрационный номер": reg_num}
        
        # Получаем все записи для данного абитуриента
        student_data = long_df[long_df["Регистрационный номер"] == reg_num]
        
        for program in programs:
            # Получаем данные по программе для данного абитуриента
            program_data = student_data[student_data["Программа"] == program]
            
            budget_priority = None
            commercial_priority = None
            scores = None
            
            for _, row in program_data.iterrows():
                if row["Тип места"] == "budget":
                    budget_priority = row["Приоритет бюджетного места"]
                    if pd.isna(scores):  # берем баллы из первой найденной записи
                        scores = row["Сумма конкурсных баллов"]
                elif row["Тип места"] == "commercial":
                    commercial_priority = row["Приоритет коммерческого места"]
                    if pd.isna(scores):  # берем баллы из первой найденной записи
                        scores = row["Сумма конкурсных баллов"]
            
            # Добавляем данные в строку
            row_data[f"{program}_budget_priority"] = budget_priority
            row_data[f"{program}_commercial_priority"] = commercial_priority
            row_data[f"{program}_scores"] = scores
        
        result_data.append(row_data)
    
    result_df = pd.DataFrame(result_data)
    
    # Создаем MultiIndex для колонок
    columns = ["Регистрационный номер"]
    
    # Добавляем колонки программ
    for program in programs:
        columns.extend([
            f"{program}_budget_priority",
            f"{program}_commercial_priority", 
            f"{program}_scores"
        ])
    
    result_df = result_df[columns]
    
    # Создаем красивые заголовки для CSV
    new_columns = ["Регистрационный номер"]
    header_row1 = [""]  # первая строка заголовков
    header_row2 = ["Регистрационный номер"]  # вторая строка заголовков
    
    for program in programs:
        new_columns.extend([
            f"{program}_Приоритет бюджетного места",
            f"{program}_Приоритет коммерческого места",
            f"{program}_Сумма конкурсных баллов"
        ])
        # Для заголовков
        header_row1.extend(["Приоритет бюджетного места", "Приоритет коммерческого места", "Сумма конкурсных баллов"])
        header_row2.extend([program, program, program])
    
    result_df.columns = new_columns
    
    # Добавляем строки заголовков в начало DataFrame
    header_df1 = pd.DataFrame([header_row1], columns=result_df.columns)
    header_df2 = pd.DataFrame([header_row2], columns=result_df.columns)
    
    final_df = pd.concat([header_df1, header_df2, result_df], ignore_index=True)
    
    return final_df


def create_proper_xlsx(wide_df: pd.DataFrame, xlsx_file: str):
    """Создает XLSX файл с многоуровневыми заголовками (программа + тип данных)."""
    try:
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Alignment, Font, Border, Side
        
        # Удаляем строки заголовков для Excel (первые две строки)
        data_df = wide_df.iloc[2:].copy()
        
        # Создаем структуру для многоуровневых заголовков
        programs = []
        level1_headers = []  # Уровень программ
        level2_headers = []  # Уровень типов данных
        
        for col in data_df.columns:
            if col == "Регистрационный номер":
                level1_headers.append("")
                level2_headers.append("Регистрационный номер")
            elif "_Приоритет бюджетного места" in col:
                program = col.replace("_Приоритет бюджетного места", "")
                if program not in programs:
                    programs.append(program)
                level1_headers.append(program)
                level2_headers.append("Приоритет бюджетного места")
            elif "_Приоритет коммерческого места" in col:
                program = col.replace("_Приоритет коммерческого места", "")
                level1_headers.append(program)
                level2_headers.append("Приоритет коммерческого места")
            elif "_Сумма конкурсных баллов" in col:
                program = col.replace("_Сумма конкурсных баллов", "")
                level1_headers.append(program)
                level2_headers.append("Сумма конкурсных баллов")
        
        # Создаем новый Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сводная таблица абитуриентов"
        
        # Записываем заголовки первого уровня (программы)
        for col_idx, header in enumerate(level1_headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Записываем заголовки второго уровня (типы данных)
        for col_idx, header in enumerate(level2_headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Записываем данные (начиная с 3-й строки)
        for row_idx, (_, row) in enumerate(data_df.iterrows(), 3):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Преобразуем значения для корректного отображения
                if pd.isna(value) or value == "":
                    cell.value = None
                elif isinstance(value, str) and value.replace('.', '').replace('-', '').isdigit():
                    try:
                        cell.value = float(value) if '.' in value else int(value)
                    except:
                        cell.value = value
                else:
                    cell.value = value
        
        # Объединяем ячейки для заголовков программ
        # Объединяем ячейки по группам (3 колонки на программу)
        col_idx = 2  # начинаем с колонки B (после регистрационного номера)
        while col_idx <= len(level1_headers):
            if col_idx <= len(level1_headers) - 2:  # убеждаемся, что есть 3 колонки
                program_name = level1_headers[col_idx - 1]
                if program_name:  # если есть название программы
                    # Сначала устанавливаем название программы
                    ws.cell(row=1, column=col_idx).value = program_name
                    # Очищаем соседние ячейки перед объединением
                    ws.cell(row=1, column=col_idx + 1).value = None
                    ws.cell(row=1, column=col_idx + 2).value = None
                    # Теперь объединяем 3 колонки
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
            col_idx += 3  # переходим к следующей программе
        
        # Настраиваем границы для заголовков
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx in range(1, len(level1_headers) + 1):
            ws.cell(row=1, column=col_idx).border = thin_border
            ws.cell(row=2, column=col_idx).border = thin_border
        
        # Автоматическая ширина колонок для первых нескольких колонок
        ws.column_dimensions['A'].width = 20  # Регистрационный номер
        for col_idx in range(2, min(20, len(level1_headers) + 1)):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 12
        
        # Закрепляем первые две строки и первый столбец
        ws.freeze_panes = 'B3'
        
        # Сохраняем файл
        wb.save(xlsx_file)
        
    except ImportError:
        print("⚠ Для создания XLSX файлов нужно установить openpyxl: pip install openpyxl")
        print("  Создан только CSV файл")
    except Exception as e:
        print(f"⚠ Ошибка при создании XLSX: {e}")
        print("  Создан только CSV файл")


# ―――――――――― CLI ――――――――――

def main():
    p = argparse.ArgumentParser(description="Собирает CSV с абитуриентами из папок budget и commercial в сводную таблицу.")
    p.add_argument("csv_root", help="Корневая папка с подпапками csv/budget и csv/commercial")
    args = p.parse_args()

    csv_root = Path(args.csv_root).expanduser().resolve()
    budget_dir = csv_root / "csv" / "budget"
    commercial_dir = csv_root / "csv" / "commercial"
    
    long_df, wide_df = build_combined_from_folders(budget_dir, commercial_dir)

    # Сохраняем CSV
    output_file = "pivot.csv"
    wide_df.to_csv(output_file, index=False, encoding='utf-8-sig')
    print(f"\n✓ Создан файл {output_file}")
    
    # Создаем также XLSX файл с правильной структурой
    xlsx_file = "pivot.xlsx"
    create_proper_xlsx(wide_df, xlsx_file)
    print(f"✓ Создан файл {xlsx_file}")
    
    print(f"\nОба файла содержат {len(wide_df)-2} строк данных")

if __name__ == "__main__":
    main()